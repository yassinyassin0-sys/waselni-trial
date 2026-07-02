#!/usr/bin/env python3
"""Seed the Waselni matching-taxonomy workbook — UAE-market edition.
Career families = UAE sectors; interests + events tuned to the UAE.
The .xlsx is the SOURCE OF TRUTH Yassin edits; xlsx-to-js.py regenerates
the app's code copy from it.

Run:  python3 seed-taxonomy-xlsx.py
"""
import colorsys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = "/Users/yassinyassin/Documents/Business/Businesses/Carpool-ParkingShare/App /Prototype/Backend/Matchmaking Algorithm/waselni-taxonomy.xlsx"

# ---- Career families = UAE sectors (ordered by local prominence) ----
FAMILIES = [
 ("Real Estate & Property", ["Real Estate Broker","Property Consultant","Leasing Manager","Off-Plan Sales Specialist","Property Manager","Mortgage Advisor","Real Estate Valuer","Real Estate Investment Advisor"]),
 ("Banking, Finance & Wealth", ["Relationship Manager","Wealth Manager","Private Banker","Investment Advisor","Portfolio Manager","Credit Analyst","Financial Analyst","Fund Manager","Accountant","Compliance Officer"]),
 ("Construction & Engineering", ["Civil Engineer","Quantity Surveyor","MEP Engineer","Structural Engineer","Site Engineer","Project Manager","Planning Engineer","Architect","Cost Estimator"]),
 ("Trade, Logistics & Supply Chain", ["Freight Forwarder","Supply Chain Manager","Logistics Coordinator","Procurement Officer","Import/Export Officer","Warehouse Manager","Commodities Trader","Customs Broker"]),
 ("Aviation & Airlines", ["Cabin Crew","Purser","First Officer","Captain","Aircraft Maintenance Engineer","Flight Dispatcher","Airport Services Agent","Air Traffic Controller"]),
 ("Hospitality & Tourism", ["Guest Relations Manager","Front Office Manager","Hotel Manager","F&B Manager","Executive Chef","Restaurant Manager","Events Manager","Concierge","Sommelier"]),
 ("Oil, Gas & Energy", ["Petroleum Engineer","Reservoir Engineer","Drilling Engineer","Process Engineer","HSE Officer","Geologist","Production Engineer","Renewable Energy Engineer"]),
 ("Government & Public Sector", ["Public Relations Officer (PRO)","Government Relations Officer","Emiratisation Officer","Policy Advisor","Regulatory Affairs Officer","Protocol Officer","Immigration Officer"]),
 ("Healthcare & Medicine", ["General Practitioner","Specialist Consultant","Registered Nurse","Surgeon","Dentist","Pharmacist","Physiotherapist","Radiologist","Healthcare Administrator"]),
 ("Consulting, Legal & Professional Services", ["Management Consultant","Strategy Consultant","Auditor","Tax Advisor","Corporate Lawyer","Legal Counsel","Risk Advisor","M&A Analyst","Actuary"]),
 ("Technology, Fintech & Web3", ["Software Engineer","Data Scientist","AI/ML Engineer","Cybersecurity Analyst","Product Manager","DevOps Engineer","Blockchain Developer","UX/UI Designer","VARA Compliance Officer"]),
 ("Media, Marketing & Creative", ["Marketing Manager","Social Media Manager","Content Creator","PR Manager","Brand Manager","Graphic Designer","Photographer","Videographer","Creative Director"]),
 ("Luxury Retail & Fashion", ["Luxury Sales Associate","Boutique Manager","Brand Ambassador","Visual Merchandiser","Client Advisor","Personal Shopper","Watch & Jewellery Specialist","Stylist"]),
]

# ---- Interests (broad UAE — inclusive of the whole market) ----
CATEGORIES = [
 ("Racket sports", ["Padel","Tennis","Squash"]),
 ("Team & spectator sports", ["Cricket","Football","Rugby"]),
 ("Fitness & wellness", ["Running / run clubs","Cycling","Gym & fitness","CrossFit","Yoga","Pilates"]),
 ("Desert & adventure", ["Desert & dune-bashing","Camping & glamping","Hiking","Skydiving"]),
 ("Water & marine", ["Yachting & sailing","Kitesurfing","Wakeboarding","Paddleboarding","Scuba diving"]),
 ("Premium & spectator sport", ["Golf","Formula 1 / Motorsport","Supercars & car culture","Horse racing & polo"]),
 ("Food, coffee & nightlife", ["Brunch","Fine dining","Specialty coffee","Beach clubs"]),
 ("Arts & culture", ["Art & galleries","Museums","Live music & concerts","Photography"]),
 ("Digital & lifestyle", ["Gaming & e-sports","Travel","Reading & book clubs"]),
]

# ---- Events (UAE categories + marquee events) ----
EVENTS = [
 ("Tech & Startup", "GITEX Global · Seamless Middle East · Expand North Star"),
 ("Finance & Fintech", "Dubai FinTech Summit · Web3 / crypto conferences"),
 ("Business & Government", "World Governments Summit · SIBOS"),
 ("Energy & Industry", "ADIPEC · WETEX · Make it in the Emirates"),
 ("Real Estate & Construction", "Cityscape Global · The Big 5"),
 ("Healthcare & Life Sciences", "Arab Health · WHX Dubai"),
 ("Aviation, Defence & Security", "Dubai Airshow · IDEX / NAVDEX · Intersec"),
 ("Food & Beverage", "Gulfood · Taste of Dubai"),
 ("Art & Design", "Art Dubai · Dubai Design Week · Abu Dhabi Art"),
 ("Luxury & Retail", "Dubai Watch Week · Dubai Shopping Festival"),
 ("Motorsport & Sport", "Abu Dhabi F1 Grand Prix · Dubai World Cup · Dubai Tennis · UAE Tour"),
 ("Music & Culture", "UNTOLD Dubai · Sole DXB · UAE National Day"),
 ("Wellness & Fitness", "Dubai Fitness Challenge (30x30) · Beautyworld"),
]

WEIGHTS = [
 ("Career vs interest blend — Balanced (default)", "65% / 35%", "How much profession matters vs shared interests, by default"),
 ("Intent — Here to network", "80% / 20%", "Driver picks this per trip: leans professional"),
 ("Intent — Here for good company", "20% / 80%", "Leans shared interests / vibe"),
 ("Exact role the driver is seeking", "1.0", "Full professional credit (driver wants 'Brokers', passenger is a Real Estate Broker)"),
 ("Same career family / sector (adjacent)", "0.5", "Partial credit — same UAE sector (see Career families tab)"),
 ("Same company", "+0.3", "Boost to the professional score"),
 ("Each mutual Waselni connection", "+0.1 (max 0.3)", "Boost to the professional score"),
 ("Each exact shared interest", "0.5 (max 1.0)", "Interest credit"),
 ("Each related interest (same group)", "0.2 (max 0.6)", "Partial interest credit (see Interest groups tab)"),
 ("Seats buffer", "+1", "Driver is shown seats + 1 candidates, to allow for declines"),
 ("Suggestion threshold (% of a role)", "30%", "A 'who to meet' target becomes a SUGGESTED chip for a role once this % of that role's users have picked it (find-your-people). Scales with the pool."),
 ("Suggestion minimum picks (floor)", "2", "Safeguard: at least this many users of a role must have picked a target before it can be suggested — stops one early pick defining the role. Set to 1 to disable."),
]

# Suggested "people to meet" — SEED / cold-start only. These are SUGGESTIONS, never
# presets; the live app amends them from real usage (e.g. if many DoPs keep picking
# 'Director', it starts suggesting Directors to DoPs). Users always choose freely.
SUGGESTED = [
 ("Real Estate & Property", ["Banking, Finance & Wealth","Construction & Engineering","Consulting, Legal & Professional Services","Luxury Retail & Fashion","Government & Public Sector"]),
 ("Banking, Finance & Wealth", ["Real Estate & Property","Consulting, Legal & Professional Services","Technology, Fintech & Web3","Trade, Logistics & Supply Chain","Government & Public Sector"]),
 ("Construction & Engineering", ["Real Estate & Property","Oil, Gas & Energy","Government & Public Sector","Consulting, Legal & Professional Services","Trade, Logistics & Supply Chain"]),
 ("Trade, Logistics & Supply Chain", ["Banking, Finance & Wealth","Government & Public Sector","Oil, Gas & Energy","Consulting, Legal & Professional Services"]),
 ("Aviation & Airlines", ["Hospitality & Tourism","Government & Public Sector","Technology, Fintech & Web3","Trade, Logistics & Supply Chain"]),
 ("Hospitality & Tourism", ["Media, Marketing & Creative","Luxury Retail & Fashion","Aviation & Airlines","Real Estate & Property"]),
 ("Oil, Gas & Energy", ["Government & Public Sector","Construction & Engineering","Banking, Finance & Wealth","Consulting, Legal & Professional Services"]),
 ("Government & Public Sector", ["Real Estate & Property","Oil, Gas & Energy","Banking, Finance & Wealth","Consulting, Legal & Professional Services","Construction & Engineering"]),
 ("Healthcare & Medicine", ["Government & Public Sector","Consulting, Legal & Professional Services","Technology, Fintech & Web3","Banking, Finance & Wealth"]),
 ("Consulting, Legal & Professional Services", ["Banking, Finance & Wealth","Real Estate & Property","Government & Public Sector","Technology, Fintech & Web3","Oil, Gas & Energy"]),
 ("Technology, Fintech & Web3", ["Banking, Finance & Wealth","Consulting, Legal & Professional Services","Media, Marketing & Creative","Real Estate & Property"]),
 ("Media, Marketing & Creative", ["Technology, Fintech & Web3","Luxury Retail & Fashion","Hospitality & Tourism","Banking, Finance & Wealth"]),
 ("Luxury Retail & Fashion", ["Media, Marketing & Creative","Hospitality & Tourism","Real Estate & Property","Banking, Finance & Wealth"]),
]
INK = "2A1545"

def hue(i, n, s=0.42, l=0.55):
    r, g, b = colorsys.hls_to_rgb((i / max(n, 1)), l, s)
    return "%02X%02X%02X" % (int(r * 255), int(g * 255), int(b * 255))

def hdr(cell, hexcolor):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=hexcolor)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def wide_sheet(ws, groups):
    n = len(groups)
    for j, (name, items) in enumerate(groups):
        hdr(ws.cell(row=1, column=j + 1, value=name), hue(j, n))
        ws.column_dimensions[get_column_letter(j + 1)].width = 26
        for i, it in enumerate(items):
            ws.cell(row=i + 2, column=j + 1, value=it)
    ws.freeze_panes = "A2"

wb = openpyxl.Workbook()

# ---- Read me ----
ws = wb.active; ws.title = "Read me"
ws.column_dimensions['A'].width = 120
lines = [
 ("Waselni — Career & Interest Compatibility  ·  UAE edition", 15, True, INK),
 ("The data behind how the app matches two people to share a ride — tuned to the UAE market (real estate, finance, energy, aviation, hospitality lead; padel, cricket, brunch, desert, F1 are the social fabric).", 11, False, "444444"),
 ("Each user sets who THEY are (their career + interests) and who they want to MEET (careers/sectors) — both chosen from the Career families list below. The app never presets who you want to meet.", 11, False, INK),
 ("", 11, False, "444444"),
 ("TABS", 12, True, INK),
 ("• Career families — 13 UAE sectors. Same sector = partial 'same-field' credit even without an exact match.", 11, False, "444444"),
 ("• Interest groups — UAE interests, grouped. Same group = partial credit.", 11, False, "444444"),
 ("• Events — UAE event categories + marquee events (for the Events screen + as networking-intent signals).", 11, False, "444444"),
 ("• Suggested people to meet — a STARTING-POINT list of who to suggest for each sector. Suggestions only (never presets), and amended by real usage: if many DoPs pick 'Director', the app starts suggesting Directors to DoPs.", 11, False, "444444"), ("• Scoring weights — the numbers that tune matching. Editable.", 11, False, "444444"),
 ("", 11, False, "444444"),
 ("HOW MATCHING USES IT", 12, True, INK),
 ("Profession and interest are each scored, then blended 65% profession / 35% interest by default (shifts with the driver's trip intent). Passengers only ever see plain reasons — 'same sector', '2 shared interests' — never numbers.", 11, False, "444444"),
 ("", 11, False, "444444"),
 ("HOW TO EDIT", 12, True, INK),
 ("These are SUGGESTIONS — users can also type their own job or interest, and it's saved to their profile. Change any cell here, move a role to another sector, add or remove items; then tell Claude — the app's lists + matching are regenerated from THIS file, the source of truth.", 11, False, "444444"),
]
for i, (txt, sz, bold, color) in enumerate(lines):
    c = ws.cell(row=i + 1, column=1, value=txt)
    c.font = Font(size=sz, bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="top")

wide_sheet(wb.create_sheet("Career families"), FAMILIES)
wide_sheet(wb.create_sheet("Interest groups"), CATEGORIES)

# ---- Events ----
we = wb.create_sheet("Events")
hdr(we.cell(row=1, column=1, value="Event category"), INK)
hdr(we.cell(row=1, column=2, value="Marquee UAE events"), INK)
we.column_dimensions['A'].width = 34; we.column_dimensions['B'].width = 74
for i, (cat, ex) in enumerate(EVENTS):
    we.cell(row=i + 2, column=1, value=cat)
    we.cell(row=i + 2, column=2, value=ex)
we.freeze_panes = "A2"

# (No audience map. "Who you want to meet" is each user's own pick of
#  careers/sectors from the Career families list — the app never presets it.)

# ---- Suggested people to meet (SEED — learns from real usage) ----
wsug = wb.create_sheet("Suggested people to meet")
hdr(wsug.cell(row=1, column=1, value="If your sector is…"), INK)
hdr(wsug.cell(row=1, column=2, value="…we'll SUGGEST meeting (seed — the app learns from real picks)"), INK)
wsug.column_dimensions['A'].width = 42; wsug.column_dimensions['B'].width = 82
for i, (sec, sugg) in enumerate(SUGGESTED):
    wsug.cell(row=i + 2, column=1, value=sec)
    wsug.cell(row=i + 2, column=2, value=" · ".join(sugg))
wsug.freeze_panes = "A2"

# ---- Scoring weights ----
wsc = wb.create_sheet("Scoring weights")
for j, h in enumerate(["Setting", "Value", "What it does"]):
    hdr(wsc.cell(row=1, column=j + 1, value=h), INK)
wsc.column_dimensions['A'].width = 44; wsc.column_dimensions['B'].width = 18; wsc.column_dimensions['C'].width = 66
for i, (a, b, c) in enumerate(WEIGHTS):
    wsc.cell(row=i + 2, column=1, value=a); wsc.cell(row=i + 2, column=2, value=b)
    cc = wsc.cell(row=i + 2, column=3, value=c); cc.alignment = Alignment(wrap_text=True, vertical="top")
wsc.freeze_panes = "A2"

wb.save(OUT)

chk = openpyxl.load_workbook(OUT)
print("saved:", OUT.split("/")[-1])
print("tabs:", chk.sheetnames)
print("sectors:", len(FAMILIES), "| job titles:", sum(len(f[1]) for f in FAMILIES))
print("interest groups:", len(CATEGORIES), "| interests:", sum(len(c[1]) for c in CATEGORIES))
print("event categories:", len(EVENTS))
